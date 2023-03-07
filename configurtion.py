import openpyxl

from google_drive_library.get import download_file_with_service_acc

email_col_name = "Email column name"
phone_col_name = "Phone number column name"
bday_col_name = "Birthday date column name"
bday_notification_advance = "Birthday notification advance"
sender_email = "Sender email"
sender_email_app_password = "Sender email app password"


def load_settings(service, files):
    # Find configuration.xlsx file
    config_file = [f for f in files if f['name'] == 'configuration.xlsx']

    # If it's not there, throw error
    if len(config_file) != 1:
        print('Excel file "configurations.xlsx" ne postoji ili nije podeljen sa servis nalogom')
        return None

    # Remove config file from files that will be searched
    files.remove(config_file[0])

    # Download config file
    file = download_file_with_service_acc(service, config_file[0]['id'])

    # Load settings from file
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.worksheets[0]

    settings = {
        email_col_name: ws["e5"].value,
        phone_col_name: ws["e6"].value,
        bday_col_name: ws["e7"].value,
        bday_notification_advance: int(ws["e8"].value),
        sender_email: ws["e9"].value,
        sender_email_app_password: ws["e10"].value,
    }

    return settings
