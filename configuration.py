import openpyxl

from drive_access.get import download_file_as_bytes

CONFIGURATION_FILE_NAME = 'configuration.xlsx'

EMAIL_COL_NAME = "Email column name"
PHONE_COL_NAME = "Phone number column name"
BDAY_COL_NAME = "Birthday date column name"
BDAY_NOTIFICATION_ADVANCE = "Birthday notification advance"
SENDER_EMAIL = "Sender email"
SENDER_EMAIL_APP_PASSWORD = "Sender email password"


class ConfigurationError(Exception):
    """Custom exception class for configuration-related errors."""
    pass


def find_configuration_file(files, file_name=CONFIGURATION_FILE_NAME):
    """
    Searches for the configuration file in the provided list of files.

    Args:
        files (list): List of files to search for the configuration file.
        file_name (str, optional): Name of the configuration file.

    Returns:
        dict: The configuration file as a dictionary containing its metadata.

    Raises:
        ConfigurationError: If the configuration file is not found or multiple files are found with the same name.
    """
    config_files = [f for f in files if f['name'] == file_name]

    if len(config_files) != 1:
        raise ConfigurationError(f'Expected exactly one configuration file "{file_name}", found {len(config_files)}')

    return config_files[0]


def download_configuration_file(service, config_file):
    """
    Downloads the configuration file using the provided service.

    Args:
        service (object): A service object to access the file storage API.
        config_file (dict): A dictionary containing the metadata of the configuration file.

    Returns:
        BytesIO: The configuration file as a bytes-like object.
    """
    return download_file_as_bytes(service, config_file['id'])


def load_settings_from_workbook(workbook):
    """
    Extracts settings from the provided workbook.

    Args:
        workbook (openpyxl.Workbook): The configuration workbook.

    Returns:
        dict: A dictionary containing the settings.
    """
    ws = workbook.worksheets[0]

    settings = {
        EMAIL_COL_NAME: ws["e5"].value,
        PHONE_COL_NAME: ws["e6"].value,
        BDAY_COL_NAME: ws["e7"].value,
        BDAY_NOTIFICATION_ADVANCE: int(ws["e8"].value),
        SENDER_EMAIL: ws["e9"].value,
        SENDER_EMAIL_APP_PASSWORD: ws["e10"].value,
    }

    return settings


def load_settings(service, files):
    """
    Loads settings from the configuration file.

    Args:
        service (object): A service object to access the file storage API.
        files (list): List of files to search for the configuration file.

    Returns:
        dict: A dictionary containing the settings.
    """
    config_file = find_configuration_file(files)
    files.remove(config_file)

    downloaded_file = download_configuration_file(service, config_file)

    wb = openpyxl.load_workbook(downloaded_file, read_only=True)

    settings = load_settings_from_workbook(wb)
    return settings
